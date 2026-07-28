from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.views.generic import TemplateView, ListView, View
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from .models import Employee, Attendance, LeaveRequest, AuditLog
from .services import calculate_monthly_salary, approve_leave, reject_leave
from django.contrib import messages
from datetime import datetime, date, timedelta
import calendar
import openpyxl


def get_or_create_employee(user):
    employee = Employee.objects.filter(user=user).first()
    if not employee:
        emp_id = f"EMP-{user.id:03d}"
        if Employee.objects.filter(employee_id=emp_id).exists():
            emp_id = f"EMP-U{user.id}"
        employee = Employee.objects.create(
            user=user,
            employee_id=emp_id,
            monthly_salary=50000
        )
    return employee


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'attendance/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['employee'] = get_or_create_employee(self.request.user)
        return context


class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'attendance/employee_list.html'
    context_object_name = 'employees'


class MarkAttendanceView(LoginRequiredMixin, View):
    def get(self, request):
        today = timezone.localtime().date()

        try:
            year = int(request.GET.get('year', today.year))
            month = int(request.GET.get('month', today.month))
        except ValueError:
            year = today.year
            month = today.month

        if month == 1:
            prev_month, prev_year = 12, year - 1
        else:
            prev_month, prev_year = month - 1, year

        if month == 12:
            next_month, next_year = 1, year + 1
        else:
            next_month, next_year = month + 1, year

        employee = get_or_create_employee(request.user)

        cal = calendar.Calendar(firstweekday=6)  # Sunday first
        month_days = cal.monthdatescalendar(year, month)

        monthly_records = []
        attendance_dict = {}
        if employee:
            qs = Attendance.objects.filter(employee=employee, date__year=year, date__month=month).order_by('date')
            monthly_records = list(qs)
            for record in qs:
                attendance_dict[record.date] = record.status

        calendar_data = []
        for week in month_days:
            week_data = []
            for d in week:
                week_data.append({
                    'date': d,
                    'day': d.day,
                    'is_current_month': d.month == month,
                    'is_today': d == today,
                    'status': attendance_dict.get(d),
                })
            calendar_data.append(week_data)

        has_checked_in = False
        has_checked_out = False
        working_hours = None

        if employee:
            try:
                today_att = Attendance.objects.get(employee=employee, date=today)
                has_checked_in = bool(today_att.time_in)
                if today_att.time_out:
                    has_checked_out = True
                    t_in = datetime.combine(date.min, today_att.time_in)
                    t_out = datetime.combine(date.min, today_att.time_out)
                    working_hours = round((t_out - t_in).total_seconds() / 3600, 2)
            except Attendance.DoesNotExist:
                pass

        leave_requests = LeaveRequest.objects.filter(employee=employee) if employee else []

        context = {
            'today': today,
            'calendar_data': calendar_data,
            'has_checked_in': has_checked_in,
            'has_checked_out': has_checked_out,
            'working_hours': working_hours,
            'employee': employee,
            'month_name': calendar.month_name[month],
            'year': year,
            'prev_month': prev_month,
            'prev_year': prev_year,
            'next_month': next_month,
            'next_year': next_year,
            'leave_requests': leave_requests,
            'monthly_records': monthly_records,
        }
        return render(request, 'attendance/mark_attendance.html', context)

    def post(self, request):
        today = timezone.localtime().date()
        action = request.POST.get('action')

        employee = get_or_create_employee(request.user)

        if action == 'checkin':
            Attendance.objects.update_or_create(
                employee=employee, date=today,
                defaults={'status': 'PRESENT', 'time_in': timezone.localtime().time()}
            )
            messages.success(request, 'Successfully checked in for today!')

        elif action == 'checkout':
            attendance = Attendance.objects.filter(employee=employee, date=today).first()
            if attendance and not attendance.time_out:
                now_time = timezone.localtime().time()
                attendance.time_out = now_time
                if attendance.time_in:
                    t_in = datetime.combine(today, attendance.time_in)
                    t_out = datetime.combine(today, now_time)
                    if t_out < t_in:
                        t_out += timedelta(days=1)
                    hours_worked = (t_out - t_in).total_seconds() / 3600.0
                    if hours_worked > 6.0:
                        attendance.status = 'PRESENT'
                    else:
                        attendance.status = 'HALF_DAY'
                attendance.save()
                messages.success(request, f'Successfully checked out. Working time: {attendance.total_working_hours} ({attendance.get_status_display()}).')

        return redirect('mark_attendance')


class ApplyLeaveView(LoginRequiredMixin, View):
    def get(self, request):
        employee = get_or_create_employee(request.user)

        leave_requests = LeaveRequest.objects.filter(employee=employee) if employee else []
        context = {
            'employee': employee,
            'leave_type_choices': LeaveRequest.LEAVE_TYPE_CHOICES,
            'leave_requests': leave_requests,
        }
        return render(request, 'attendance/apply_leave.html', context)

    def post(self, request):
        employee = get_or_create_employee(request.user)

        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        leave_type = request.POST.get('leave_type')
        reason = request.POST.get('reason', '').strip()

        if not (start_date_str and end_date_str and leave_type and reason):
            messages.error(request, 'All fields are required.')
            return redirect('apply_leave')

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid date format.')
            return redirect('apply_leave')

        if end_date < start_date:
            messages.error(request, 'End date cannot be before start date.')
            return redirect('apply_leave')

        LeaveRequest.objects.create(
            employee=employee,
            start_date=start_date,
            end_date=end_date,
            leave_type=leave_type,
            reason=reason,
        )
        messages.success(request, 'Leave request submitted successfully! Pending admin approval.')
        return redirect('apply_leave')


class SalaryView(LoginRequiredMixin, View):
    def get(self, request):
        today = timezone.now().date()

        try:
            year = int(request.GET.get('year', today.year))
            month = int(request.GET.get('month', today.month))
        except ValueError:
            year, month = today.year, today.month

        if month == 1:
            prev_month, prev_year = 12, year - 1
        else:
            prev_month, prev_year = month - 1, year

        if month == 12:
            next_month, next_year = 1, year + 1
        else:
            next_month, next_year = month + 1, year

        employee = get_or_create_employee(request.user)

        salary_data = None
        monthly_records = []
        if employee:
            salary_data = calculate_monthly_salary(employee, year, month)
            monthly_records = list(
                Attendance.objects.filter(employee=employee, date__year=year, date__month=month)
                .order_by('date')
            )

        context = {
            'today': today,
            'employee': employee,
            'salary_data': salary_data,
            'monthly_records': monthly_records,
            'year': year,
            'month': month,
            'prev_month': prev_month,
            'prev_year': prev_year,
            'next_month': next_month,
            'next_year': next_year,
        }
        return render(request, 'attendance/salary.html', context)


class AttendanceReportView(LoginRequiredMixin, ListView):
    model = Attendance
    template_name = 'attendance/reports.html'
    context_object_name = 'attendance_records'
    paginate_by = 100

    def get_queryset(self):
        queryset = Attendance.objects.select_related('employee', 'employee__user').all().order_by('-date', 'employee__employee_id')
        
        if not (self.request.user.is_staff or self.request.user.is_superuser):
            queryset = queryset.filter(employee__user=self.request.user)

        report_type = self.request.GET.get('report_type', 'daily')
        today = timezone.localtime().date()

        if report_type == 'daily':
            selected_date_str = self.request.GET.get('date', today.strftime('%Y-%m-%d'))
            try:
                selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(date=selected_date)
            except ValueError:
                queryset = queryset.filter(date=today)
        else:
            month = self.request.GET.get('month', today.strftime('%Y-%m'))
            try:
                year, m = month.split('-')
                queryset = queryset.filter(date__year=year, date__month=m)
            except ValueError:
                queryset = queryset.filter(date__year=today.year, date__month=today.month)

        employee_id = self.request.GET.get('employee')
        if employee_id and (self.request.user.is_staff or self.request.user.is_superuser):
            queryset = queryset.filter(employee__id=employee_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['employees'] = Employee.objects.select_related('user').all().order_by('employee_id')
        today = timezone.localtime().date()
        
        report_type = self.request.GET.get('report_type', 'daily')
        context['report_type'] = report_type
        context['selected_date'] = self.request.GET.get('date', today.strftime('%Y-%m-%d'))
        context['selected_month'] = self.request.GET.get('month', today.strftime('%Y-%m'))
        context['selected_employee'] = self.request.GET.get('employee', '')
        
        qs = self.get_queryset()
        context['total_records'] = qs.count()
        context['present_count'] = qs.filter(status='PRESENT').count()
        context['absent_count'] = qs.filter(status='ABSENT').count()
        context['leave_count'] = qs.filter(status='LEAVE').count()
        context['half_day_count'] = qs.filter(status='HALF_DAY').count()
        return context


class ExportAttendanceReportView(LoginRequiredMixin, View):
    def get(self, request):
        queryset = Attendance.objects.select_related('employee', 'employee__user').all().order_by('-date', 'employee__employee_id')
        
        if not (request.user.is_staff or request.user.is_superuser):
            queryset = queryset.filter(employee__user=request.user)

        report_type = request.GET.get('report_type', 'daily')
        today = timezone.localtime().date()

        if report_type == 'daily':
            selected_date_str = request.GET.get('date', today.strftime('%Y-%m-%d'))
            try:
                selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(date=selected_date)
                period_label = selected_date.strftime('%Y-%m-%d')
            except ValueError:
                queryset = queryset.filter(date=today)
                period_label = today.strftime('%Y-%m-%d')
            sheet_title = f"Daily Report ({period_label})"
            filename = f"Daily_Attendance_Report_{period_label}.xlsx"
        else:
            month = request.GET.get('month', today.strftime('%Y-%m'))
            try:
                year, m = month.split('-')
                queryset = queryset.filter(date__year=year, date__month=m)
                period_label = month
            except ValueError:
                queryset = queryset.filter(date__year=today.year, date__month=today.month)
                period_label = today.strftime('%Y-%m')
            sheet_title = f"Monthly Report ({period_label})"
            filename = f"Monthly_Attendance_Report_{period_label}.xlsx"

        employee_id = request.GET.get('employee')
        if employee_id and (request.user.is_staff or request.user.is_superuser):
            queryset = queryset.filter(employee__id=employee_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]

        headers = [
            'Sl No', 'Date', 'Employee ID', 'Employee Name', 'Department',
            'Designation', 'Status', 'Check In', 'Check Out', 'Total Working Hours'
        ]
        ws.append(headers)

        for idx, rec in enumerate(queryset, start=1):
            emp_name = rec.employee.user.get_full_name() or rec.employee.user.username
            time_in_str = rec.time_in.strftime('%H:%M') if rec.time_in else '-'
            time_out_str = rec.time_out.strftime('%H:%M') if rec.time_out else '-'

            row = [
                idx,
                rec.date.strftime('%d/%m/%Y'),
                rec.employee.employee_id,
                emp_name,
                rec.employee.department or '',
                rec.employee.designation or '',
                rec.get_status_display(),
                time_in_str,
                time_out_str,
                rec.total_working_hours
            ]
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ManageLeavesView(LoginRequiredMixin, View):
    def get(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('dashboard')

        status_filter = request.GET.get('status', 'ALL')
        
        queryset = LeaveRequest.objects.select_related('employee', 'employee__user').all()
        
        pending_count = queryset.filter(status='PENDING').count()
        approved_count = queryset.filter(status='APPROVED').count()
        rejected_count = queryset.filter(status='REJECTED').count()
        total_count = queryset.count()

        if status_filter != 'ALL':
            leave_requests = queryset.filter(status=status_filter)
        else:
            leave_requests = queryset

        context = {
            'leave_requests': leave_requests,
            'status_filter': status_filter,
            'pending_count': pending_count,
            'approved_count': approved_count,
            'rejected_count': rejected_count,
            'total_count': total_count,
        }
        return render(request, 'attendance/manage_leaves.html', context)

    def post(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('dashboard')

        leave_id = request.POST.get('leave_id')
        action = request.POST.get('action')

        try:
            leave_request = LeaveRequest.objects.get(id=leave_id)
            if action == 'approve':
                approve_leave(leave_request, reviewed_by=request.user)
                messages.success(request, f'Leave for {leave_request.employee} approved successfully! Attendance updated.')
            elif action == 'reject':
                reject_leave(leave_request, reviewed_by=request.user)
                messages.warning(request, f'Leave for {leave_request.employee} has been rejected.')
        except LeaveRequest.DoesNotExist:
            messages.error(request, 'Leave request not found.')

        status_filter = request.POST.get('status_filter', 'ALL')
        return redirect(f"/leave/manage/?status={status_filter}")


class CreateEmployeeView(LoginRequiredMixin, View):
    def get(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('dashboard')
        return render(request, 'attendance/create_employee.html')

    def post(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('dashboard')

        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password', '').strip()
        employee_id = request.POST.get('employee_id', '').strip()
        phone_number = request.POST.get('phone_number', '').strip()
        address = request.POST.get('address', '').strip()
        department = request.POST.get('department', 'General').strip()
        designation = request.POST.get('designation', 'Staff').strip()
        
        # Extended fields
        hiring_date_str = request.POST.get('date_of_hiring', '').strip()
        pt_location = request.POST.get('pt_location', '').strip()
        annual_ctc_str = request.POST.get('annual_ctc', '0').strip()
        pan_number = request.POST.get('pan_number', '').strip()
        provident_fund = request.POST.get('provident_fund') == 'on' or request.POST.get('provident_fund') == 'Y'
        pf_uan = request.POST.get('pf_uan', '').strip()
        esic_number = request.POST.get('esic_number', '').strip()
        gender = request.POST.get('gender', '').strip()
        dob_str = request.POST.get('date_of_birth', '').strip()
        bank_account = request.POST.get('bank_account_number', '').strip()
        bank_ifsc = request.POST.get('bank_ifsc_code', '').strip()

        if not (first_name and email and password):
            messages.error(request, 'First name, email, and password are required.')
            return render(request, 'attendance/create_employee.html', {'request_data': request.POST})

        if User.objects.filter(username=email).exists() or User.objects.filter(email=email).exists():
            messages.error(request, f'A user with email/username "{email}" already exists.')
            return render(request, 'attendance/create_employee.html', {'request_data': request.POST})

        try:
            annual_ctc = float(annual_ctc_str) if annual_ctc_str else 0.0
        except ValueError:
            annual_ctc = 0.0

        monthly_salary = round(annual_ctc / 12.0, 2) if annual_ctc > 0 else 30000.00

        date_of_hiring = None
        if hiring_date_str:
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    date_of_hiring = datetime.strptime(hiring_date_str, fmt).date()
                    break
                except ValueError:
                    pass

        date_of_birth = None
        if dob_str:
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    date_of_birth = datetime.strptime(dob_str, fmt).date()
                    break
                except ValueError:
                    pass

        user = User.objects.create_user(
            username=email,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name
        )

        emp_id = employee_id
        if not emp_id:
            next_id = User.objects.count()
            emp_id = f"EMP-{next_id:03d}"
            while Employee.objects.filter(employee_id=emp_id).exists():
                next_id += 1
                emp_id = f"EMP-{next_id:03d}"

        Employee.objects.create(
            user=user,
            employee_id=emp_id,
            department=department if department else 'General',
            designation=designation if designation else 'Staff',
            phone_number=phone_number,
            address=address,
            monthly_salary=monthly_salary,
            annual_ctc=annual_ctc,
            date_of_hiring=date_of_hiring,
            pt_location=pt_location,
            pan_number=pan_number,
            provident_fund=provident_fund,
            pf_uan=pf_uan,
            esic_number=esic_number,
            gender=gender,
            date_of_birth=date_of_birth,
            bank_account_number=bank_account,
            bank_ifsc_code=bank_ifsc
        )

        messages.success(request, f'Employee account for {first_name} {last_name} ({email}) created successfully! ID: {emp_id}')
        return redirect('create_employee')


class BulkUploadEmployeeView(LoginRequiredMixin, View):
    def post(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('dashboard')

        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Please select an Excel file (.xlsx) to upload.')
            return redirect('create_employee')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file format. Please upload a valid Excel spreadsheet (.xlsx).')
            return redirect('create_employee')

        default_password = request.POST.get('default_password', 'Emp@12345').strip() or 'Emp@12345'

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            if 'Employee Data' in wb.sheetnames:
                sheet = wb['Employee Data']

            # Locate header row (search first 5 rows)
            header_row_idx = None
            header_map = {}
            for r in range(1, min(6, sheet.max_row + 1)):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                row_str = " ".join([str(v).lower() for v in row_vals if v is not None])
                if 'email' in row_str or 'employee id' in row_str or 'name' in row_str:
                    header_row_idx = r
                    for col_idx, cell_val in enumerate(row_vals, start=1):
                        if cell_val:
                            key = str(cell_val).strip().lower()
                            header_map[key] = col_idx
                    break

            if not header_row_idx:
                messages.error(request, 'Could not detect header row in Excel sheet. Ensure headers like "Email", "Name", "Employee ID" exist.')
                return redirect('create_employee')

            def get_col_val(row_idx, keywords):
                if isinstance(keywords, str):
                    keywords = [keywords]
                for kw in keywords:
                    for h_key, col_idx in header_map.items():
                        if kw.lower() in h_key:
                            return sheet.cell(row=row_idx, column=col_idx).value
                return None

            def parse_date_val(val):
                if not val:
                    return None
                if isinstance(val, (datetime, date)):
                    return val if isinstance(val, date) else val.date()
                val_str = str(val).strip()
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
                    try:
                        return datetime.strptime(val_str, fmt).date()
                    except ValueError:
                        pass
                return None

            def parse_decimal_val(val, default=0.0):
                if val is None or val == '':
                    return default
                try:
                    clean_str = str(val).replace(',', '').replace('$', '').replace('₹', '').strip()
                    return float(clean_str)
                except (ValueError, TypeError):
                    return default

            created_count = 0
            updated_count = 0
            errors = []

            with transaction.atomic():
                for r in range(header_row_idx + 1, sheet.max_row + 1):
                    email_raw = get_col_val(r, ['email'])
                    name_raw = get_col_val(r, ['name'])

                    if not email_raw and not name_raw:
                        continue

                    email = str(email_raw).strip().lower() if email_raw else ''
                    name = str(name_raw).strip() if name_raw else ''

                    if not email or '@' not in email:
                        errors.append(f"Row {r}: Skipped due to missing or invalid email ('{email_raw}').")
                        continue

                    emp_id_raw = get_col_val(r, ['employee id', 'emp id'])
                    emp_id = str(emp_id_raw).strip() if emp_id_raw else ''

                    name_parts = name.split(' ', 1)
                    first_name = name_parts[0] if name_parts else 'Employee'
                    last_name = name_parts[1] if len(name_parts) > 1 else ''

                    user = User.objects.filter(username=email).first() or User.objects.filter(email=email).first()
                    if not user:
                        user = User.objects.create_user(
                            username=email,
                            email=email,
                            password=default_password,
                            first_name=first_name,
                            last_name=last_name
                        )
                        user_created = True
                    else:
                        user_created = False
                        user.first_name = first_name or user.first_name
                        user.last_name = last_name or user.last_name
                        user.save()

                    if not emp_id:
                        next_num = User.objects.count()
                        emp_id = f"TSTPL-{next_num:04d}"
                        while Employee.objects.filter(employee_id=emp_id).exists():
                            next_num += 1
                            emp_id = f"TSTPL-{next_num:04d}"

                    title_raw = get_col_val(r, ['title', 'designation'])
                    dept_raw = get_col_val(r, ['department'])
                    phone_raw = get_col_val(r, ['phone'])
                    hiring_date_raw = get_col_val(r, ['hiring', 'date of hiring'])
                    pt_location_raw = get_col_val(r, ['professional tax location', 'pt location', 'state'])
                    annual_ctc_raw = get_col_val(r, ['annual ctc', 'ctc'])
                    pan_raw = get_col_val(r, ['pan'])
                    pf_raw = get_col_val(r, ['provident fund'])
                    pf_uan_raw = get_col_val(r, ['pf uan', 'uan'])
                    esic_raw = get_col_val(r, ['esic'])
                    gender_raw = get_col_val(r, ['gender'])
                    dob_raw = get_col_val(r, ['date of birth', 'dob'])
                    bank_acc_raw = get_col_val(r, ['bank account', 'account number'])
                    bank_ifsc_raw = get_col_val(r, ['bank ifsc', 'ifsc'])
                    taxable_curr_raw = get_col_val(r, ['taxable salary paid in current'])
                    exemptions_curr_raw = get_col_val(r, ['exemptions in current'])
                    tds_curr_raw = get_col_val(r, ['tds deducted in current'])
                    past_taxable_raw = get_col_val(r, ['past taxable salary'])
                    past_tds_raw = get_col_val(r, ['past tds'])

                    annual_ctc = parse_decimal_val(annual_ctc_raw, 0.0)
                    monthly_salary = round(annual_ctc / 12.0, 2) if annual_ctc > 0 else 30000.00
                    provident_fund = str(pf_raw).strip().upper() in ('Y', 'YES', 'TRUE', '1') if pf_raw else False

                    # Safe Employee lookup & ID collision prevention
                    employee = Employee.objects.filter(user=user).first()
                    if not employee and emp_id:
                        employee = Employee.objects.filter(employee_id=emp_id).first()

                    if not emp_id:
                        next_num = User.objects.count()
                        emp_id = f"TSTPL-{next_num:04d}"
                        while Employee.objects.filter(employee_id=emp_id).exists():
                            next_num += 1
                            emp_id = f"TSTPL-{next_num:04d}"

                    if not employee:
                        while Employee.objects.filter(employee_id=emp_id).exists():
                            emp_id = f"{emp_id}-DUP"
                        employee = Employee.objects.create(
                            user=user,
                            employee_id=emp_id
                        )
                        emp_created = True
                    else:
                        emp_created = False
                        if employee.employee_id != emp_id and emp_id:
                            if not Employee.objects.filter(employee_id=emp_id).exclude(id=employee.id).exists():
                                employee.employee_id = emp_id

                    employee.department = str(dept_raw).strip() if dept_raw else 'General'
                    employee.designation = str(title_raw).strip() if title_raw else 'Staff'
                    employee.phone_number = str(phone_raw).strip() if phone_raw else None
                    employee.annual_ctc = annual_ctc
                    employee.monthly_salary = monthly_salary
                    employee.date_of_hiring = parse_date_val(hiring_date_raw)
                    employee.pt_location = str(pt_location_raw).strip() if pt_location_raw else None
                    employee.pan_number = str(pan_raw).strip() if pan_raw else None
                    employee.provident_fund = provident_fund
                    employee.pf_uan = str(pf_uan_raw).strip() if pf_uan_raw else None
                    employee.esic_number = str(esic_raw).strip() if esic_raw else None
                    employee.gender = str(gender_raw).strip() if gender_raw else None
                    employee.date_of_birth = parse_date_val(dob_raw)
                    employee.bank_account_number = str(bank_acc_raw).strip() if bank_acc_raw else None
                    employee.bank_ifsc_code = str(bank_ifsc_raw).strip() if bank_ifsc_raw else None
                    employee.taxable_salary_current_fy = parse_decimal_val(taxable_curr_raw)
                    employee.exemptions_current_fy = parse_decimal_val(exemptions_curr_raw)
                    employee.tds_deducted_current_fy = parse_decimal_val(tds_curr_raw)
                    employee.past_taxable_salary = parse_decimal_val(past_taxable_raw)
                    employee.past_tds = parse_decimal_val(past_tds_raw)
                    employee.save()

                    if user_created or emp_created:
                        created_count += 1
                    else:
                        updated_count += 1

            if created_count > 0 or updated_count > 0:
                messages.success(request, f"Excel import successful! {created_count} employee(s) created, {updated_count} updated. Default password for new logins: '{default_password}'.")
            if errors:
                for err in errors[:5]:
                    messages.warning(request, err)

        except Exception as e:
            messages.error(request, f"Error processing Excel file: {str(e)}")

        return redirect('create_employee')



class ChangePasswordView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'attendance/change_password.html')

    def post(self, request):
        current_password = request.POST.get('current_password', '')
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')

        if not current_password or not new_password or not confirm_password:
            messages.error(request, 'All password fields are required.')
            return render(request, 'attendance/change_password.html')

        if not request.user.check_password(current_password):
            messages.error(request, 'Your current password is incorrect.')
            return render(request, 'attendance/change_password.html')

        if new_password != confirm_password:
            messages.error(request, 'The new password and confirmation password do not match.')
            return render(request, 'attendance/change_password.html')

        if len(new_password) < 8:
            messages.error(request, 'New password must be at least 8 characters long.')
            return render(request, 'attendance/change_password.html')

        if current_password == new_password:
            messages.error(request, 'New password cannot be the same as your current password.')
            return render(request, 'attendance/change_password.html')

        user = request.user
        user.set_password(new_password)
        user.save()
        update_session_auth_hash(request, user)

        messages.success(request, 'Your password has been changed successfully!')
        return redirect('change_password')


class AdminChangeUserPasswordView(LoginRequiredMixin, View):
    def post(self, request, user_id):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('dashboard')

        new_password = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()

        if not new_password or not confirm_password:
            messages.error(request, 'Password fields cannot be empty.')
            return redirect('employee_list')

        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
            return redirect('employee_list')

        if len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
            return redirect('employee_list')

        try:
            target_user = User.objects.get(id=user_id)
            target_user.set_password(new_password)
            target_user.save()

            AuditLog.objects.create(
                actor=request.user,
                action_type='RESET_PASSWORD',
                target_user=target_user,
                details=f"Admin '{request.user.username}' reset password for employee account '{target_user.email}'."
            )
            messages.success(request, f'Password for {target_user.get_full_name() or target_user.username} updated successfully!')
        except User.DoesNotExist:
            messages.error(request, 'Target user not found.')

        return redirect('employee_list')


class ToggleUserActiveView(LoginRequiredMixin, View):
    def post(self, request, user_id):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('employee_list')

        target_user = get_object_or_404(User, id=user_id)
        if target_user == request.user:
            messages.error(request, 'You cannot deactivate your own administrative account.')
            return redirect('employee_list')

        target_user.is_active = not target_user.is_active
        target_user.save()

        employee = Employee.objects.filter(user=target_user).first()
        if employee:
            if not target_user.is_active:
                exit_date_str = request.POST.get('date_of_exit')
                if exit_date_str:
                    try:
                        employee.date_of_exit = datetime.strptime(exit_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        employee.date_of_exit = timezone.now().date()
                else:
                    employee.date_of_exit = timezone.now().date()
            else:
                employee.date_of_exit = None
            employee.save()

        action_code = 'REACTIVATE' if target_user.is_active else 'DEACTIVATE'
        details_txt = "Account reactivated and departure status cleared." if target_user.is_active else f"Account deactivated. Marked exit date as {employee.date_of_exit if employee else 'today'}."
        AuditLog.objects.create(
            actor=request.user,
            action_type=action_code,
            target_user=target_user,
            details=details_txt
        )

        status_label = "deactivated (marked as left)" if not target_user.is_active else "reactivated"
        messages.success(request, f"Employee {target_user.get_full_name() or target_user.username} account has been {status_label}.")
        return redirect('employee_list')


class EditEmployeeView(LoginRequiredMixin, View):
    def post(self, request, employee_id):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('employee_list')

        employee = get_object_or_404(Employee, id=employee_id)
        user = employee.user

        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()

        emp_id_input = request.POST.get('employee_id', '').strip()
        department = request.POST.get('department', '').strip()
        designation = request.POST.get('designation', '').strip()
        phone_number = request.POST.get('phone_number', '').strip()
        address = request.POST.get('address', '').strip()

        pt_location = request.POST.get('pt_location', '').strip()
        pan_number = request.POST.get('pan_number', '').strip()
        provident_fund = request.POST.get('provident_fund') in ('true', 'True', 'on', '1')
        pf_uan = request.POST.get('pf_uan', '').strip()
        esic_number = request.POST.get('esic_number', '').strip()
        gender = request.POST.get('gender', '').strip()
        bank_account_number = request.POST.get('bank_account_number', '').strip()
        bank_ifsc_code = request.POST.get('bank_ifsc_code', '').strip()

        date_of_hiring_str = request.POST.get('date_of_hiring', '').strip()
        date_of_birth_str = request.POST.get('date_of_birth', '').strip()
        date_of_exit_str = request.POST.get('date_of_exit', '').strip()

        annual_ctc_str = request.POST.get('annual_ctc', '').strip()
        monthly_salary_str = request.POST.get('monthly_salary', '').strip()
        taxable_fy_str = request.POST.get('taxable_salary_current_fy', '').strip()
        exemptions_fy_str = request.POST.get('exemptions_current_fy', '').strip()
        tds_fy_str = request.POST.get('tds_deducted_current_fy', '').strip()

        if emp_id_input and emp_id_input != employee.employee_id:
            if Employee.objects.filter(employee_id=emp_id_input).exclude(id=employee.id).exists():
                messages.error(request, f"Employee ID '{emp_id_input}' is already in use by another employee.")
                return redirect('employee_list')
            employee.employee_id = emp_id_input

        if first_name:
            user.first_name = first_name
        if last_name:
            user.last_name = last_name
        if email and email != user.email:
            if User.objects.filter(email=email).exclude(id=user.id).exists():
                messages.error(request, f"Email address '{email}' is already registered to another user.")
                return redirect('employee_list')
            user.email = email
            user.username = email
        user.save()

        def parse_date(d_str):
            if not d_str:
                return None
            try:
                return datetime.strptime(d_str, '%Y-%m-%d').date()
            except ValueError:
                return None

        def parse_dec(v_str, default=0.0):
            if not v_str:
                return default
            try:
                return float(v_str)
            except ValueError:
                return default

        employee.department = department or 'General'
        employee.designation = designation or 'Staff'
        employee.phone_number = phone_number or None
        employee.address = address or None

        annual_ctc = parse_dec(annual_ctc_str, 0.0)
        employee.annual_ctc = annual_ctc
        if monthly_salary_str:
            employee.monthly_salary = parse_dec(monthly_salary_str, 0.0)
        elif annual_ctc > 0:
            employee.monthly_salary = round(annual_ctc / 12.0, 2)

        employee.pt_location = pt_location or None
        employee.pan_number = pan_number or None
        employee.provident_fund = provident_fund
        employee.pf_uan = pf_uan or None
        employee.esic_number = esic_number or None
        employee.gender = gender or None
        employee.bank_account_number = bank_account_number or None
        employee.bank_ifsc_code = bank_ifsc_code or None

        employee.date_of_hiring = parse_date(date_of_hiring_str) if date_of_hiring_str else employee.date_of_hiring
        employee.date_of_birth = parse_date(date_of_birth_str) if date_of_birth_str else employee.date_of_birth
        
        if 'date_of_exit' in request.POST:
            employee.date_of_exit = parse_date(date_of_exit_str)

        employee.taxable_salary_current_fy = parse_dec(taxable_fy_str)
        employee.exemptions_current_fy = parse_dec(exemptions_fy_str)
        employee.tds_deducted_current_fy = parse_dec(tds_fy_str)

        employee.save()

        AuditLog.objects.create(
            actor=request.user,
            action_type='EDIT_PROFILE',
            target_user=user,
            details=f"Admin '{request.user.username}' edited employee profile details for '{user.get_full_name() or user.username}'."
        )

        messages.success(request, f"Employee '{user.get_full_name() or user.username}' updated successfully.")
        return redirect('employee_list')


class ExportEmployeesExcelView(LoginRequiredMixin, View):
    def get(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('employee_list')

        queryset = Employee.objects.select_related('user').all().order_by('id')
        q = request.GET.get('q', '').strip()
        status = request.GET.get('status', '').strip()

        if q:
            queryset = queryset.filter(
                Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q) |
                Q(user__email__icontains=q) |
                Q(user__username__icontains=q) |
                Q(employee_id__icontains=q) |
                Q(department__icontains=q) |
                Q(designation__icontains=q)
            )
        if status == 'active':
            queryset = queryset.filter(user__is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(user__is_active=False)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees Directory"

        headers = [
            'Sl No', 'Employee ID', 'Full Name', 'Email', 'Department', 'Designation',
            'Status', 'Date of Hiring', 'Date of Exit', 'Monthly Salary (INR)', 'Annual CTC (INR)',
            'PAN Number', 'PF Account', 'Bank Account', 'IFSC Code'
        ]
        ws.append(headers)

        for idx, emp in enumerate(queryset, start=1):
            full_name = emp.user.get_full_name() or emp.user.username
            status_str = 'Active' if emp.user.is_active else 'Left Company'
            hiring_date = emp.date_of_hiring.strftime('%d/%m/%Y') if emp.date_of_hiring else ''
            exit_date = emp.date_of_exit.strftime('%d/%m/%Y') if emp.date_of_exit else ''
            pf_str = 'Yes' if emp.provident_fund else 'No'

            row = [
                idx, emp.employee_id, full_name, emp.user.email, emp.department or '', emp.designation or '',
                status_str, hiring_date, exit_date, float(emp.monthly_salary or 0), float(emp.annual_ctc or 0),
                emp.pan_number or '', pf_str, emp.bank_account_number or '', emp.bank_ifsc_code or ''
            ]
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Employee_Directory_Export.xlsx"'
        wb.save(response)
        return response


class AuditLogView(LoginRequiredMixin, ListView):
    model = AuditLog
    template_name = 'attendance/audit_log.html'
    context_object_name = 'audit_logs'
    paginate_by = 25

    def get_queryset(self):
        if not (self.request.user.is_staff or self.request.user.is_superuser):
            return AuditLog.objects.none()

        queryset = AuditLog.objects.select_related('actor', 'target_user').all()
        q = self.request.GET.get('q', '').strip()
        action = self.request.GET.get('action', '').strip()

        if q:
            queryset = queryset.filter(
                Q(actor__username__icontains=q) |
                Q(actor__first_name__icontains=q) |
                Q(target_user__username__icontains=q) |
                Q(target_user__first_name__icontains=q) |
                Q(details__icontains=q)
            )

        if action:
            queryset = queryset.filter(action_type=action)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get('q', '')
        context['action'] = self.request.GET.get('action', '')
        context['action_choices'] = AuditLog.ACTION_CHOICES
        return context


class EmployeeProfileView(LoginRequiredMixin, View):
    def get(self, request):
        employee = get_or_create_employee(request.user)
        today = timezone.localtime().date()
        first_day = today.replace(day=1)
        attendance_records = Attendance.objects.filter(
            employee=employee,
            date__gte=first_day,
            date__lte=today
        ).order_by('-date')

        present_count = attendance_records.filter(status='PRESENT').count()
        absent_count = attendance_records.filter(status='ABSENT').count()
        leave_count = attendance_records.filter(status='LEAVE').count()
        half_day_count = attendance_records.filter(status='HALF_DAY').count()

        context = {
            'employee': employee,
            'attendance_records': attendance_records[:10],
            'present_count': present_count,
            'absent_count': absent_count,
            'leave_count': leave_count,
            'half_day_count': half_day_count,
        }
        return render(request, 'attendance/employee_profile.html', context)


class FnFSettlementView(LoginRequiredMixin, View):
    def get(self, request, employee_id):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('employee_list')

        employee = get_object_or_404(Employee, id=employee_id)
        exit_date = employee.date_of_exit or timezone.now().date()
        days_in_exit_month = calendar.monthrange(exit_date.year, exit_date.month)[1]
        worked_days = exit_date.day

        monthly_salary = float(employee.monthly_salary or 0)
        daily_rate = monthly_salary / days_in_exit_month if days_in_exit_month else 0
        pro_rated_salary = daily_rate * worked_days

        annual_ctc = float(employee.annual_ctc or 0)
        tds_estimate = pro_rated_salary * 0.05 if annual_ctc > 500000 else 0
        net_payable = max(0, pro_rated_salary - tds_estimate)

        context = {
            'employee': employee,
            'exit_date': exit_date,
            'days_in_exit_month': days_in_exit_month,
            'worked_days': worked_days,
            'monthly_salary': monthly_salary,
            'daily_rate': round(daily_rate, 2),
            'pro_rated_salary': round(pro_rated_salary, 2),
            'tds_estimate': round(tds_estimate, 2),
            'net_payable': round(net_payable, 2),
        }
        return render(request, 'attendance/fnf_settlement.html', context)






